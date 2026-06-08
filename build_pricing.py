"""
NutriLens Fiyatlandırma Modeli Excel Oluşturucu
Çalıştır: python3 build_pricing.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

import os

wb = Workbook()

# ── Renkler ──────────────────────────────────────────────────────────────────
BG_DARK     = "0D1117"
BG_CARD     = "161B22"
BG_GREEN    = "1A4A2E"
BG_RED      = "4A1A1A"
BG_YELLOW   = "4A3A00"
BG_BLUE     = "0D2033"
C_GREEN     = "58A6FF"
C_RED       = "F87171"
C_YELLOW    = "FCD34D"
C_WHITE     = "E6EDF3"
C_SUBTEXT   = "8B949E"
C_HEADER    = "21262D"
C_LIGHT     = "F0F6FF"
C_ORANGE    = "FB923C"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin(sides="all"):
    s = Side(style="thin", color="30363D")
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    elif sides == "bottom":
        return Border(bottom=s)
    return Border()

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_cell(ws, row, col, text, bg=C_HEADER, fg=C_WHITE,
                bold=True, size=11, h="center", wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(h=h, wrap=wrap)
    c.border = border_thin()
    return c

def data_cell(ws, row, col, value, bg=BG_CARD, fg=C_WHITE,
              bold=False, h="center", fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg)
    c.alignment = align(h=h, wrap=wrap)
    c.border = border_thin()
    if fmt:
        c.number_format = fmt
    return c

def section_title(ws, row, col, text, span_end, bg=BG_BLUE, fg=C_GREEN):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = Font(bold=True, color=fg, size=12, name="Calibri")
    c.alignment = align(h="left")
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=span_end)
    c.border = border_thin()
    return c

# ════════════════════════════════════════════════════════════════════════════
#  SAYFA 1 — API TOKEN MALİYETLERİ
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📡 API Maliyetleri"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A4"

# Başlık
ws1.merge_cells("A1:I1")
t = ws1["A1"]
t.value = "🔬 NutriLens — API Token Maliyet Hesaplama Tablosu"
t.fill = fill(BG_DARK)
t.font = Font(bold=True, color=C_GREEN, size=16, name="Calibri")
t.alignment = align()

ws1.merge_cells("A2:I2")
t2 = ws1["A2"]
t2.value = "Tüm sarı hücreler değiştirilebilir · Model: claude-haiku-4-5 · Fiyatlar: USD/1M token"
t2.fill = fill(BG_DARK)
t2.font = Font(color=C_SUBTEXT, size=10, name="Calibri")
t2.alignment = align()
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 18

# ── Model Fiyat Tablosu ───────────────────────────────────────────────────
r = 4
section_title(ws1, r, 1, "⚙️  MODEL FİYATLARI (USD / 1M token)  ← Değiştirebilirsin", 9)
r += 1

headers = ["Model", "Input $/1M", "Output $/1M", "Vision Input $/1M", "Görüntü Flat Fee $/img", "", "", "", ""]
for i, h in enumerate(headers, 1):
    c = header_cell(ws1, r, i, h, bg=C_HEADER)
r += 1

model_data = [
    ["claude-haiku-4-5", 0.80, 4.00, 1.20, 0.0048, "", "", "", ""],
    ["claude-sonnet-4-6 (alternatif)", 3.00, 15.00, 3.00, 0.0048, "", "", "", ""],
]

model_rows = {}
for model in model_data:
    for col_i, val in enumerate(model, 1):
        if col_i <= 5 and isinstance(val, (int, float)) and col_i > 1:
            c = data_cell(ws1, r, col_i, val, bg=BG_YELLOW, fg="000000",
                          fmt='"$"#,##0.0000')
            c.fill = PatternFill("solid", fgColor="FCD34D")
        elif col_i == 1:
            data_cell(ws1, r, col_i, val, h="left", fg=C_WHITE)
        else:
            data_cell(ws1, r, col_i, val)
    model_rows[model[0]] = r
    r += 1

# Named ranges (manual references for formulas)
HAIKU_INPUT   = f"B{model_rows['claude-haiku-4-5']}"  # $/1M input
HAIKU_OUTPUT  = f"C{model_rows['claude-haiku-4-5']}"  # $/1M output
HAIKU_VIS_IN  = f"D{model_rows['claude-haiku-4-5']}"  # $/1M vision input
HAIKU_VIS_FLAT= f"E{model_rows['claude-haiku-4-5']}"  # $/image flat

r += 1
# ── Özellik Bazlı Token Analizi ───────────────────────────────────────────
section_title(ws1, r, 1, "🧮  ÖZELLİK BAZLI TOKEN KULLANIMI VE MALİYET (tek çağrı başına)", 9)
r += 1

feat_headers = [
    "Özellik", "Input Token\n(tahmini)", "Output Token\n(tahmini)",
    "Görüntü var?", "Tek Çağrı\nInput ($)", "Tek Çağrı\nOutput ($)",
    "Görüntü ($)", "TOPLAM\n$/çağrı", "Açıklama"
]
for i, h in enumerate(feat_headers, 1):
    header_cell(ws1, r, i, h, wrap=True)
ws1.row_dimensions[r].height = 32
r += 1

# Token tahminleri — gerçek prompt uzunluklarına göre
features = [
    # [ad, input_tok, output_tok, has_img, açıklama]
    ["📸 Görsel Analiz (kamera)",    1800, 350, True,
     "Prompt ~1400tok + resim metaveri, yanıt JSON ~350tok"],
    ["📝 Anlatarak Analiz (metin)",  900,  350, False,
     "Metin prompt ~900tok, yanıt JSON ~350tok"],
    ["💬 Koç Mesajı (tek tur)",      650,  180, False,
     "System ~250tok + geçmiş 10 msg avg + kullanıcı sorusu + yanıt"],
    ["🔄 Koç Oturumu (10 mesaj)",    3500, 1800, False,
     "Her mesajda geçmiş büyür; 10 turlu oturum toplamı"],
    ["📊 Stok Analizi (çağrı)",      400,  200, False,
     "Conflict detection servisi — çakışma kontrolü"],
]

feat_start_row = r
for feat in features:
    ad, inp, out, img, acik = feat
    data_cell(ws1, r, 1, ad, h="left", fg=C_WHITE)
    # Sarı — değiştirilebilir
    c_in = ws1.cell(row=r, column=2, value=inp)
    c_in.fill = PatternFill("solid", fgColor="FCD34D")
    c_in.font = Font(bold=False, color="000000", name="Calibri")
    c_in.alignment = align()
    c_in.border = border_thin()
    c_in.number_format = "#,##0"

    c_out = ws1.cell(row=r, column=3, value=out)
    c_out.fill = PatternFill("solid", fgColor="FCD34D")
    c_out.font = Font(bold=False, color="000000", name="Calibri")
    c_out.alignment = align()
    c_out.border = border_thin()
    c_out.number_format = "#,##0"

    data_cell(ws1, r, 4, "Evet" if img else "Hayır",
              bg=(BG_GREEN if img else BG_CARD))

    # Maliyet formülleri
    col_in_ref  = f"B{r}"
    col_out_ref = f"C{r}"
    has_img_f   = f"D{r}"

    # Input cost = (token / 1M) * (vision_rate if img else normal_rate)
    if img:
        cost_in_formula = f"=({col_in_ref}/1000000)*{HAIKU_VIS_IN}"
    else:
        cost_in_formula = f"=({col_in_ref}/1000000)*{HAIKU_INPUT}"
    cost_out_formula = f"=({col_out_ref}/1000000)*{HAIKU_OUTPUT}"
    cost_img_formula = f"=IF(D{r}=\"Evet\",{HAIKU_VIS_FLAT},0)"
    cost_total_f     = f"=E{r}+F{r}+G{r}"

    for col_f, fmt_f, formula in [
        (5, '"$"#,##0.000000', cost_in_formula),
        (6, '"$"#,##0.000000', cost_out_formula),
        (7, '"$"#,##0.000000', cost_img_formula),
        (8, '"$"#,##0.000000', cost_total_f),
    ]:
        c = ws1.cell(row=r, column=col_f, value=formula)
        c.fill = fill(BG_DARK)
        c.font = Font(color=C_GREEN, name="Calibri", bold=(col_f==8))
        c.alignment = align()
        c.border = border_thin()
        c.number_format = fmt_f

    data_cell(ws1, r, 9, acik, h="left", fg=C_SUBTEXT, wrap=True)
    ws1.row_dimensions[r].height = 28
    r += 1

r += 1
# ── Türk Lirası Dönüşüm ───────────────────────────────────────────────────
section_title(ws1, r, 1, "💱  KUR DÖNÜŞÜMÜ", 9)
r += 1
header_cell(ws1, r, 1, "USD/TRY Kuru")
header_cell(ws1, r, 2, "Değer")
for i in range(3, 10):
    data_cell(ws1, r, i, "")
r += 1
kur_row = r
c_kur = ws1.cell(row=r, column=2, value=38.5)
c_kur.fill = PatternFill("solid", fgColor="FCD34D")
c_kur.font = Font(bold=True, color="000000", name="Calibri")
c_kur.alignment = align()
c_kur.border = border_thin()
c_kur.number_format = '#,##0.00 "₺"'

data_cell(ws1, r, 1, "USD/TRY", h="left")
for i in range(3, 10):
    data_cell(ws1, r, i, "")
r += 2

# Sütun genişlikleri
col_widths_1 = [28, 14, 14, 12, 14, 14, 12, 14, 38]
for i, w in enumerate(col_widths_1, 1):
    set_col_width(ws1, i, w)


# ════════════════════════════════════════════════════════════════════════════
#  SAYFA 2 — KULLANICI PROFİLLERİ
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("👤 Kullanıcı Profilleri")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A5"

ws2.merge_cells("A1:M1")
t = ws2["A1"]
t.value = "👤 NutriLens — Kullanıcı Profili & Aylık Kullanım Tahminleri"
t.fill = fill(BG_DARK)
t.font = Font(bold=True, color=C_GREEN, size=16, name="Calibri")
t.alignment = align()

ws2.merge_cells("A2:M2")
t2 = ws2["A2"]
t2.value = "Sarı hücreler = değiştirilebilir. Hesaplama otomatik güncellenir."
t2.fill = fill(BG_DARK)
t2.font = Font(color=C_SUBTEXT, size=10, name="Calibri")
t2.alignment = align()
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 18

r2 = 4

# Profil başlıkları
section_title(ws2, r2, 1,
    "📊  KULLANICI PROFİLLERİ — Aylık Kullanım Tahminleri  ← Sarı hücreleri değiştir",
    13)
r2 += 1

profiles = ["Düşük Kullanıcı", "Ortalama Kullanıcı",
            "Çok Kullanan", "Fazla Kullanan"]
profile_cols = [2, 5, 8, 11]  # Her profil 3 sütun

# Profil renkleri
prof_colors = ["1A3A4A", "1A4A2E", "4A3A00", "4A1A1A"]
prof_fg     = ["60A5FA", "7EE787", "FCD34D", "F87171"]

sub_cols = ["Değer", "Açıklama", ""]

for pi, (pname, pcol) in enumerate(zip(profiles, profile_cols)):
    ws2.merge_cells(start_row=r2, start_column=pcol,
                    end_row=r2, end_column=pcol+1)
    c = ws2.cell(row=r2, column=pcol, value=pname)
    c.fill = fill(prof_colors[pi])
    c.font = Font(bold=True, color=prof_fg[pi], size=11, name="Calibri")
    c.alignment = align()
    c.border = border_thin()
    ws2.cell(row=r2, column=pcol+2, value="").fill = fill(BG_DARK)

header_cell(ws2, r2, 1, "Özellik / Metrik", bg=C_HEADER)
r2 += 1

# Sub-headers
header_cell(ws2, r2, 1, "", bg=C_HEADER)
for pcol in profile_cols:
    header_cell(ws2, r2, pcol,   "Adet/Ay", bg=C_HEADER, size=9)
    header_cell(ws2, r2, pcol+1, "Maliyet $", bg=C_HEADER, size=9)
    ws2.cell(row=r2, column=pcol+2, value="").fill = fill(BG_DARK)
r2 += 1

# Kullanım metrikleri ve varsayılan değerler
metrics = [
    # (label, [düşük, ortalama, çok, fazla])
    ("📸 Görsel Fotoğraf Analizi/ay",     [5,   15,  35,  60]),
    ("📝 Anlatarak Metin Analizi/ay",     [3,   10,  20,  40]),
    ("💬 Koç Mesajı (tek)/ay",            [5,   20,  60, 150]),
    ("🔄 Koç Oturumu Uzun (10 msg)/ay",   [0,    2,   5,  12]),
    ("📊 Stok/Çakışma Analizi/ay",        [2,    8,  20,  40]),
    ("🔍 USDA Barcode Araması/ay",        [3,   10,  20,  35]),
    ("🍱 Edamam Besin Verisi/ay",         [2,    8,  18,  30]),
]

metric_rows = {}
for mdata in metrics:
    label, defaults = mdata
    data_cell(ws2, r2, 1, label, h="left", fg=C_WHITE)
    for pi, (pcol, default_val) in enumerate(zip(profile_cols, defaults)):
        # Sarı — değiştirilebilir
        c = ws2.cell(row=r2, column=pcol, value=default_val)
        c.fill = PatternFill("solid", fgColor="FCD34D")
        c.font = Font(color="000000", name="Calibri")
        c.alignment = align()
        c.border = border_thin()
        c.number_format = "#,##0"
        # Maliyet sütunu — formülle doldurulacak (Sayfa 3'te birleşik hesap)
        data_cell(ws2, r2, pcol+1, "→ Sayfa 3", bg=BG_DARK, fg=C_SUBTEXT)
        ws2.cell(row=r2, column=pcol+2, value="").fill = fill(BG_DARK)
    metric_rows[label] = r2
    r2 += 1

r2 += 1
section_title(ws2, r2, 1,
    "💰  BEKLENEN AYLIK GELİR (Abonelik)", 13)
r2 += 1
header_cell(ws2, r2, 1, "Fiyat Tipi", bg=C_HEADER)
for pi, (pname, pcol) in enumerate(zip(profiles, profile_cols)):
    ws2.merge_cells(start_row=r2, start_column=pcol,
                    end_row=r2, end_column=pcol+1)
    header_cell(ws2, r2, pcol, pname, bg=prof_colors[pi], fg=prof_fg[pi])
    ws2.cell(row=r2, column=pcol+2, value="").fill = fill(BG_DARK)
r2 += 1
plan_price_row = r2
data_cell(ws2, r2, 1, "📱 Aylık Plan Fiyatı (₺)", h="left")
for pcol in profile_cols:
    c = ws2.cell(row=r2, column=pcol, value=149)
    c.fill = PatternFill("solid", fgColor="FCD34D")
    c.font = Font(bold=True, color="000000", name="Calibri", size=12)
    c.alignment = align()
    c.border = border_thin()
    c.number_format = '#,##0.00 "₺"'
    ws2.cell(row=r2, column=pcol+2, value="").fill = fill(BG_DARK)

col_widths_2 = [28] + [11, 11, 3] * 4
for i, w in enumerate(col_widths_2, 1):
    set_col_width(ws2, i, w)


# ════════════════════════════════════════════════════════════════════════════
#  SAYFA 3 — MALİYET & KAR/ZARAR ANALİZİ
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("💰 Maliyet & Kar-Zarar")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A5"

ws3.merge_cells("A1:L1")
t = ws3["A1"]
t.value = "💰 NutriLens — Aylık Maliyet, Gelir ve Kar/Zarar Analizi"
t.fill = fill(BG_DARK)
t.font = Font(bold=True, color=C_GREEN, size=16, name="Calibri")
t.alignment = align()

ws3.merge_cells("A2:L2")
t2 = ws3["A2"]
t2.value = "Tüm hesaplamalar '📡 API Maliyetleri' ve '👤 Kullanıcı Profilleri' sayfalarından otomatik çekiliyor"
t2.fill = fill(BG_DARK)
t2.font = Font(color=C_SUBTEXT, size=10, name="Calibri")
t2.alignment = align()
ws3.row_dimensions[1].height = 30

r3 = 4

# Profil tanımları (renk referansı)
profiles_3 = ["Düşük", "Ortalama", "Çok", "Fazla"]
prof_cols_3 = [3, 5, 7, 9]  # B=açıklama, sonra 4 profil çifti

section_title(ws3, r3, 1, "📊  KULLANICI BAŞINA AYLIK MALİYET (USD)", 11)
r3 += 1

# Header row
header_cell(ws3, r3, 1, "Özellik")
header_cell(ws3, r3, 2, "Birim Maliyet\n($/çağrı)", wrap=True)
for pi, (pname, pcol) in enumerate(zip(profiles_3, prof_cols_3)):
    ws3.merge_cells(start_row=r3, start_column=pcol,
                    end_row=r3, end_column=pcol+1)
    c = header_cell(ws3, r3, pcol, f"{pname}\n(adet | $)", wrap=True,
                    bg=prof_colors[pi], fg=prof_fg[pi])
r3 += 1

# Yardımcı map: özellik adı → sayfa1 toplam maliyet hücresi
# Sayfa 1'deki özellik satırları feat_start_row'dan başlıyor
# Biz formülle referans edeceğiz (cross-sheet)
feat_labels_short = [
    "Görsel Analiz",
    "Metin Analiz",
    "Koç Tekil Msg",
    "Koç Uzun Oturum",
    "Stok Analizi",
]

# Sayfa 1'deki toplam maliyet sütunu = H (col 8)
# Sayfa 1 özellik satırları feat_start_row+0..+4
s1_cost_rows = list(range(feat_start_row, feat_start_row + 5))

# Sayfa 2 kullanım sayıları
# metric_rows: label → row_number
# Sütunlar: pcol = kullanım adedi
s2_metric_labels_in_order = [
    "📸 Görsel Fotoğraf Analizi/ay",
    "📝 Anlatarak Metin Analizi/ay",
    "💬 Koç Mesajı (tek)/ay",
    "🔄 Koç Oturumu Uzun (10 msg)/ay",
    "📊 Stok/Çakışma Analizi/ay",
]
s2_profile_cols = [2, 5, 8, 11]  # sayfa2'deki kullanım adet sütunları

cost_row_refs = {}  # feat_idx → row in ws3
for fi, (flabel, s1r) in enumerate(zip(feat_labels_short, s1_cost_rows)):
    s2_label = s2_metric_labels_in_order[fi]
    s2_row = metric_rows[s2_label]

    data_cell(ws3, r3, 1, flabel, h="left")
    # Birim maliyet formülü → Sayfa1!H<row>
    c_unit = ws3.cell(row=r3, column=2,
                      value=f"='📡 API Maliyetleri'!H{s1r}")
    c_unit.fill = fill(BG_DARK)
    c_unit.font = Font(color=C_YELLOW, name="Calibri")
    c_unit.alignment = align()
    c_unit.border = border_thin()
    c_unit.number_format = '"$"#,##0.000000'

    for pi, (pcol, s2_pcol) in enumerate(zip(prof_cols_3, s2_profile_cols)):
        # Adet
        c_qty = ws3.cell(row=r3, column=pcol,
                         value=f"='👤 Kullanıcı Profilleri'!{get_column_letter(s2_pcol)}{s2_row}")
        c_qty.fill = fill(BG_DARK)
        c_qty.font = Font(color=C_WHITE, name="Calibri")
        c_qty.alignment = align()
        c_qty.border = border_thin()
        c_qty.number_format = "#,##0"

        # Maliyet
        c_cost = ws3.cell(row=r3, column=pcol+1,
                          value=f"=B{r3}*{get_column_letter(pcol)}{r3}")
        c_cost.fill = fill(BG_DARK)
        c_cost.font = Font(color=C_GREEN, name="Calibri")
        c_cost.alignment = align()
        c_cost.border = border_thin()
        c_cost.number_format = '"$"#,##0.000000'

    cost_row_refs[fi] = r3
    r3 += 1

# USDA/Edamam — flat API ücretsiz ama kota var; gösterim için 0 maliyet
for extra_label in ["USDA Barcode (ücretsiz)", "Edamam (ücretsiz tier)"]:
    data_cell(ws3, r3, 1, extra_label, h="left", fg=C_SUBTEXT)
    for pcol in range(2, 11):
        data_cell(ws3, r3, pcol, 0 if pcol != 1 else "", fg=C_SUBTEXT)
    r3 += 1

# TOPLAM AYLIK MALİYET PER KULLANICI
r3 += 1
section_title(ws3, r3, 1, "🔢  KULLANICI BAŞINA TOPLAM AYLIK MALİYET (USD)", 11)
r3 += 1
total_cost_row = r3
data_cell(ws3, r3, 1, "Toplam Maliyet/Kullanıcı/Ay (USD)", h="left",
          bold=True, fg=C_WHITE)
data_cell(ws3, r3, 2, "")

sum_cost_rows = list(cost_row_refs.values())
for pi, pcol in enumerate(prof_cols_3):
    cost_col = get_column_letter(pcol + 1)
    sum_formula = "=SUM(" + ",".join(
        [f"{cost_col}{rr}" for rr in sum_cost_rows]) + ")"
    c = ws3.cell(row=r3, column=pcol+1, value=sum_formula)
    c.fill = fill(prof_colors[pi])
    c.font = Font(bold=True, color=prof_fg[pi], size=12, name="Calibri")
    c.alignment = align()
    c.border = border_thin()
    c.number_format = '"$"#,##0.0000'
    ws3.cell(row=r3, column=pcol, value="↑ sum").fill = fill(prof_colors[pi])
    ws3.cell(row=r3, column=pcol).font = Font(color=prof_fg[pi], size=9,
                                               name="Calibri", italic=True)
    ws3.cell(row=r3, column=pcol).alignment = align()
    ws3.cell(row=r3, column=pcol).border = border_thin()

r3 += 1
# TRY karşılığı
kur_ref = f"'📡 API Maliyetleri'!B{kur_row}"
data_cell(ws3, r3, 1, "Toplam Maliyet/Kullanıcı/Ay (₺)", h="left",
          bold=True, fg=C_WHITE)
data_cell(ws3, r3, 2, "")
for pi, pcol in enumerate(prof_cols_3):
    cost_col_ref = f"{get_column_letter(pcol+1)}{total_cost_row}"
    c = ws3.cell(row=r3, column=pcol+1,
                 value=f"={cost_col_ref}*{kur_ref}")
    c.fill = fill(prof_colors[pi])
    c.font = Font(bold=True, color=prof_fg[pi], size=12, name="Calibri")
    c.alignment = align()
    c.border = border_thin()
    c.number_format = '#,##0.00 "₺"'
    ws3.cell(row=r3, column=pcol, value="").fill = fill(prof_colors[pi])
    ws3.cell(row=r3, column=pcol).border = border_thin()
try_cost_row = r3

r3 += 2

# ── KAR/ZARAR ANALİZİ ────────────────────────────────────────────────────
section_title(ws3, r3, 1,
    "📈  KAR / ZARAR ANALİZİ — Aylık abonelik fiyatına göre", 11)
r3 += 1

# Fiyat senaryoları (₺)
price_scenarios = [79, 99, 119, 149, 179, 199, 249, 299]
plan_label_col = 1
header_cell(ws3, r3, plan_label_col, "Aylık Plan Fiyatı (₺)")
header_cell(ws3, r3, 2, "")
for pi, (pname, pcol) in enumerate(zip(profiles_3, prof_cols_3)):
    ws3.merge_cells(start_row=r3, start_column=pcol,
                    end_row=r3, end_column=pcol+1)
    header_cell(ws3, r3, pcol, f"Kar/Zarar {pname} (₺)",
                bg=prof_colors[pi], fg=prof_fg[pi])
r3 += 1

pnl_start = r3
for price in price_scenarios:
    c_price = ws3.cell(row=r3, column=1, value=price)
    c_price.fill = PatternFill("solid", fgColor="FCD34D")
    c_price.font = Font(bold=True, color="000000", size=12, name="Calibri")
    c_price.alignment = align()
    c_price.border = border_thin()
    c_price.number_format = '#,##0 "₺"'
    data_cell(ws3, r3, 2, "")

    for pi, pcol in enumerate(prof_cols_3):
        cost_try_ref = f"{get_column_letter(pcol+1)}{try_cost_row}"
        # Kar = Fiyat - Maliyet
        formula = f"=A{r3}-{cost_try_ref}"
        c = ws3.cell(row=r3, column=pcol+1, value=formula)
        c.fill = fill(BG_DARK)
        c.font = Font(bold=True, color=C_WHITE, size=11, name="Calibri")
        c.alignment = align()
        c.border = border_thin()
        c.number_format = '#,##0.00 "₺"'
        ws3.cell(row=r3, column=pcol, value="").fill = fill(BG_DARK)
        ws3.cell(row=r3, column=pcol).border = border_thin()
    r3 += 1

# Conditional formatting: yeşil = pozitif, kırmızı = negatif
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill as PF

green_fill = PF("solid", fgColor="1A4A2E")
red_fill   = PF("solid", fgColor="4A1A1A")
green_font = Font(bold=True, color="7EE787", name="Calibri")
red_font   = Font(bold=True, color="F87171", name="Calibri")

pnl_end = r3 - 1
for pi, pcol in enumerate(prof_cols_3):
    col_letter = get_column_letter(pcol+1)
    rng = f"{col_letter}{pnl_start}:{col_letter}{pnl_end}"
    ws3.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=green_fill, font=green_font))
    ws3.conditional_formatting.add(rng, CellIsRule(
        operator="lessThanOrEqual", formula=["0"],
        fill=red_fill, font=red_font))

r3 += 2

# ── TOPLU KULLANICI ANALİZİ ───────────────────────────────────────────────
section_title(ws3, r3, 1,
    "👥  TOPLU KULLANICI ANALİZİ — Aylık gelir ve gider (X kullanıcı)", 11)
r3 += 1
header_cell(ws3, r3, 1, "Profil")
header_cell(ws3, r3, 2, "Kullanıcı\nSayısı", wrap=True)
header_cell(ws3, r3, 3, "Aylık Plan\nFiyatı (₺)", wrap=True)
header_cell(ws3, r3, 4, "Toplam\nGelir (₺)", wrap=True)
header_cell(ws3, r3, 5, "Maliyet/\nKullanıcı (₺)", wrap=True)
header_cell(ws3, r3, 6, "Toplam\nMaliyet (₺)", wrap=True)
header_cell(ws3, r3, 7, "Net Kar/\nZarar (₺)", wrap=True)
header_cell(ws3, r3, 8, "Marj %", wrap=True)
for c in range(9, 12):
    data_cell(ws3, r3, c, "")
ws3.row_dimensions[r3].height = 32
r3 += 1

bulk_start = r3
for pi, (pname, pcol, prof_color, prof_fg_c) in enumerate(
        zip(profiles_3, prof_cols_3, prof_colors, prof_fg)):

    c_label = ws3.cell(row=r3, column=1, value=pname)
    c_label.fill = fill(prof_color)
    c_label.font = Font(bold=True, color=prof_fg_c, name="Calibri")
    c_label.alignment = align()
    c_label.border = border_thin()

    # Kullanıcı sayısı — sarı, değiştirilebilir
    default_users = [200, 500, 150, 50][pi]
    c_users = ws3.cell(row=r3, column=2, value=default_users)
    c_users.fill = PatternFill("solid", fgColor="FCD34D")
    c_users.font = Font(bold=True, color="000000", name="Calibri")
    c_users.alignment = align()
    c_users.border = border_thin()
    c_users.number_format = "#,##0"

    # Plan fiyatı — sayfa2'den referans
    s2_plan_row = plan_price_row
    s2_plan_col = get_column_letter(s2_profile_cols[pi])
    c_plan = ws3.cell(row=r3, column=3,
                      value=f"='👤 Kullanıcı Profilleri'!{s2_plan_col}{s2_plan_row}")
    c_plan.fill = fill(BG_DARK)
    c_plan.font = Font(color=C_YELLOW, name="Calibri", bold=True)
    c_plan.alignment = align()
    c_plan.border = border_thin()
    c_plan.number_format = '#,##0.00 "₺"'

    rev_formula = f"=B{r3}*C{r3}"
    c_rev = ws3.cell(row=r3, column=4, value=rev_formula)
    c_rev.fill = fill(BG_GREEN)
    c_rev.font = Font(bold=True, color="7EE787", name="Calibri")
    c_rev.alignment = align()
    c_rev.border = border_thin()
    c_rev.number_format = '#,##0.00 "₺"'

    cost_try_col = get_column_letter(pcol+1)
    c_cost_pu = ws3.cell(row=r3, column=5,
                          value=f"={cost_try_col}{try_cost_row}")
    c_cost_pu.fill = fill(BG_DARK)
    c_cost_pu.font = Font(color=C_RED, name="Calibri")
    c_cost_pu.alignment = align()
    c_cost_pu.border = border_thin()
    c_cost_pu.number_format = '#,##0.00 "₺"'

    c_total_cost = ws3.cell(row=r3, column=6, value=f"=B{r3}*E{r3}")
    c_total_cost.fill = fill(BG_RED)
    c_total_cost.font = Font(bold=True, color=C_RED, name="Calibri")
    c_total_cost.alignment = align()
    c_total_cost.border = border_thin()
    c_total_cost.number_format = '#,##0.00 "₺"'

    c_net = ws3.cell(row=r3, column=7, value=f"=D{r3}-F{r3}")
    c_net.fill = fill(BG_DARK)
    c_net.font = Font(bold=True, color=C_WHITE, size=12, name="Calibri")
    c_net.alignment = align()
    c_net.border = border_thin()
    c_net.number_format = '#,##0.00 "₺"'

    c_marj = ws3.cell(row=r3, column=8,
                       value=f"=IF(D{r3}>0,(G{r3}/D{r3})*100,0)")
    c_marj.fill = fill(BG_DARK)
    c_marj.font = Font(bold=True, color=C_WHITE, name="Calibri")
    c_marj.alignment = align()
    c_marj.border = border_thin()
    c_marj.number_format = '#,##0.0"%"'

    for c in range(9, 12):
        data_cell(ws3, r3, c, "")

    r3 += 1

bulk_end = r3 - 1

# Toplam satırı
ws3.merge_cells(f"A{r3}:B{r3}")
c_tot = ws3.cell(row=r3, column=1, value="TOPLAM")
c_tot.fill = fill(C_HEADER)
c_tot.font = Font(bold=True, color=C_WHITE, size=12, name="Calibri")
c_tot.alignment = align()
c_tot.border = border_thin()

for col_i, fmt in [(4, '#,##0.00 "₺"'), (6, '#,##0.00 "₺"'), (7, '#,##0.00 "₺"')]:
    col_l = get_column_letter(col_i)
    c = ws3.cell(row=r3, column=col_i,
                 value=f"=SUM({col_l}{bulk_start}:{col_l}{bulk_end})")
    c.fill = fill(C_HEADER)
    c.font = Font(bold=True, color=C_GREEN, size=12, name="Calibri")
    c.alignment = align()
    c.border = border_thin()
    c.number_format = fmt

for col_i in [3, 5, 8]:
    data_cell(ws3, r3, col_i, "", bg=C_HEADER)
for col_i in range(9, 12):
    data_cell(ws3, r3, col_i, "")

# CFormats
ws3.conditional_formatting.add(
    f"G{bulk_start}:G{bulk_end}",
    CellIsRule(operator="greaterThan", formula=["0"],
               fill=green_fill, font=green_font))
ws3.conditional_formatting.add(
    f"G{bulk_start}:G{bulk_end}",
    CellIsRule(operator="lessThanOrEqual", formula=["0"],
               fill=red_fill, font=red_font))

col_widths_3 = [20, 12, 14, 14, 16, 14, 14, 10, 3, 3, 3]
for i, w in enumerate(col_widths_3, 1):
    set_col_width(ws3, i, w)


# ════════════════════════════════════════════════════════════════════════════
#  SAYFA 4 — ÖZET DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📋 Özet Dashboard")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:H1")
t = ws4["A1"]
t.value = "📋 NutriLens — Fiyatlandırma Özet Dashboard"
t.fill = fill(BG_DARK)
t.font = Font(bold=True, color=C_GREEN, size=18, name="Calibri")
t.alignment = align()
ws4.row_dimensions[1].height = 36

ws4.merge_cells("A2:H2")
t2 = ws4["A2"]
t2.value = "Bu sayfa diğer sayfaları özetler. Değiştirilen herhangi bir sarı hücre burayı otomatik günceller."
t2.fill = fill(BG_DARK)
t2.font = Font(color=C_SUBTEXT, size=11, name="Calibri")
t2.alignment = align()
ws4.row_dimensions[2].height = 20

r4 = 4
for prow in range(r4, r4 + 40):
    for pcol in range(1, 9):
        ws4.cell(row=prow, column=pcol).fill = fill(BG_DARK)

# KPI Kartları
section_title(ws4, r4, 1, "⚡  ANLAM ÇIKART — API Maliyet Özeti (Tek Çağrı Başına)", 8)
r4 += 1

kpi_labels = [
    ("📸 Görsel Analiz", f"='📡 API Maliyetleri'!H{feat_start_row}"),
    ("📝 Metin Analiz",  f"='📡 API Maliyetleri'!H{feat_start_row+1}"),
    ("💬 Koç Msg",       f"='📡 API Maliyetleri'!H{feat_start_row+2}"),
    ("🔄 Koç 10-msg",    f"='📡 API Maliyetleri'!H{feat_start_row+3}"),
    ("📊 Stok Analiz",   f"='📡 API Maliyetleri'!H{feat_start_row+4}"),
]

kpi_colors = ["1A3A4A", "1A4A2E", "4A3A00", "4A1A1A", "2A1A4A"]
kpi_fgs    = ["60A5FA", "7EE787", "FCD34D", "F87171", "BF5AF2"]

for ki, (klabel, kform) in enumerate(kpi_labels):
    col = ki + 1
    if col > 8:
        r4 += 2
        col = 1

    c_lbl = ws4.cell(row=r4, column=col, value=klabel)
    c_lbl.fill = fill(kpi_colors[ki % len(kpi_colors)])
    c_lbl.font = Font(bold=True, color=kpi_fgs[ki % len(kpi_fgs)],
                      size=10, name="Calibri")
    c_lbl.alignment = align()
    c_lbl.border = border_thin()
    ws4.row_dimensions[r4].height = 22

    r4 += 1
    c_val = ws4.cell(row=r4, column=col, value=kform)
    c_val.fill = fill(kpi_colors[ki % len(kpi_colors)])
    c_val.font = Font(bold=True, color=kpi_fgs[ki % len(kpi_fgs)],
                      size=14, name="Calibri")
    c_val.alignment = align()
    c_val.border = border_thin()
    c_val.number_format = '"$"#,##0.000000'
    ws4.row_dimensions[r4].height = 28
    r4 -= 1

r4 += 3

# Kar/Zarar Özeti — seçili fiyata göre
section_title(ws4, r4, 1,
    "💡  HIZLI KAR/ZARAR — Seçili Plan Fiyatına Göre (₺)", 8)
r4 += 1

header_cell(ws4, r4, 1, "Profil")
header_cell(ws4, r4, 2, "Kullanıcı\nSayısı", wrap=True)
header_cell(ws4, r4, 3, "Plan\nFiyatı (₺)", wrap=True)
header_cell(ws4, r4, 4, "Net Kar/\nZarar (₺)", wrap=True)
header_cell(ws4, r4, 5, "Marj %", wrap=True)
header_cell(ws4, r4, 6, "Toplam\nGelir (₺)", wrap=True)
header_cell(ws4, r4, 7, "Toplam\nMaliyet (₺)", wrap=True)
header_cell(ws4, r4, 8, "Başa Baş\nFiyat (₺)", wrap=True)
ws4.row_dimensions[r4].height = 32
r4 += 1

kpi_pnl_start = r4
for pi, (pname, prof_color, prof_fg_c) in enumerate(
        zip(profiles_3, prof_colors, prof_fg)):

    c_lbl = ws4.cell(row=r4, column=1, value=pname)
    c_lbl.fill = fill(prof_color)
    c_lbl.font = Font(bold=True, color=prof_fg_c, name="Calibri")
    c_lbl.alignment = align()
    c_lbl.border = border_thin()

    # Buraya sayfa 3'teki bulk satırından bağlantı verelim
    bulk_row = bulk_start + pi
    cols_ws3 = ["B", "C", "D", "E", "F", "G", "H"]
    dest_cols = [2, 3, 7, None, 6, 4, 5]  # kullanıcı, fiyat, net, -, maliyet, gelir, marj
    mappings = {
        2: "B",   # kullanıcı sayısı
        3: "C",   # plan fiyatı
        4: "G",   # net kar/zarar
        5: "H",   # marj
        6: "D",   # toplam gelir
        7: "F",   # toplam maliyet
    }
    for dest_col, src_col in mappings.items():
        c = ws4.cell(row=r4, column=dest_col,
                     value=f"='💰 Maliyet & Kar-Zarar'!{src_col}{bulk_row}")
        c.fill = fill(prof_color if dest_col == 4 else BG_DARK)
        c.font = Font(bold=(dest_col in [4, 5]),
                      color=prof_fg_c if dest_col == 4 else C_WHITE,
                      size=(12 if dest_col == 4 else 11),
                      name="Calibri")
        c.alignment = align()
        c.border = border_thin()
        if dest_col == 3:
            c.number_format = '#,##0.00 "₺"'
        elif dest_col in [4, 6, 7]:
            c.number_format = '#,##0.00 "₺"'
        elif dest_col == 5:
            c.number_format = '#,##0.0"%"'

    # Başa baş fiyatı = maliyet/kullanıcı (TRY)
    pcol_ws3 = prof_cols_3[pi]
    c_bbe = ws4.cell(row=r4, column=8,
                     value=f"='💰 Maliyet & Kar-Zarar'!{get_column_letter(pcol_ws3+1)}{try_cost_row}")
    c_bbe.fill = fill(BG_YELLOW)
    c_bbe.font = Font(bold=True, color=C_YELLOW, name="Calibri")
    c_bbe.alignment = align()
    c_bbe.border = border_thin()
    c_bbe.number_format = '#,##0.00 "₺"'
    ws4.row_dimensions[r4].height = 26
    r4 += 1

# Toplam
ws4.merge_cells(f"A{r4}:B{r4}")
c_tot = ws4.cell(row=r4, column=1, value="GENEL TOPLAM")
c_tot.fill = fill(C_HEADER)
c_tot.font = Font(bold=True, color=C_WHITE, size=12, name="Calibri")
c_tot.alignment = align()
c_tot.border = border_thin()

for col_i, fmt in [(4, '#,##0.00 "₺"'), (6, '#,##0.00 "₺"'), (7, '#,##0.00 "₺"')]:
    col_l = get_column_letter(col_i)
    c = ws4.cell(row=r4, column=col_i,
                 value=f"=SUM({col_l}{kpi_pnl_start}:{col_l}{r4-1})")
    c.fill = fill(C_HEADER)
    c.font = Font(bold=True, color=C_GREEN, size=12, name="Calibri")
    c.alignment = align()
    c.border = border_thin()
    c.number_format = fmt
for col_i in [3, 5, 8]:
    data_cell(ws4, r4, col_i, "", bg=C_HEADER)

r4 += 2

# Notlar
section_title(ws4, r4, 1, "📝  NOTLAR & VARSAYIMLAR", 8)
r4 += 1
notes = [
    "• claude-haiku-4-5: $0.80/1M input · $4.00/1M output · Görüntü input $1.20/1M + $0.0048/img",
    "• Görsel analiz = ~1800 input + 350 output token + 1 görüntü",
    "• Metin analizi = ~900 input + 350 output token",
    "• Koç mesajı = ~650 input (sistem prompt ~250 + geçmiş + soru) + 180 output token",
    "• Koç uzun oturum (10 msg) = kümülatif ~3500 input + 1800 output token",
    "• USDA & Edamam ücretsiz tier kullanılıyor → maliyet $0",
    "• Firebase, App Store / Play Store komisyonu (%15-30) bu tabloya dahil DEĞİL",
    "• Tüm sarı hücreler değiştirilebilir. Değiştirince tüm hesaplamalar otomatik güncellenir.",
]
for note in notes:
    ws4.merge_cells(f"A{r4}:H{r4}")
    c = ws4.cell(row=r4, column=1, value=note)
    c.fill = fill(BG_DARK)
    c.font = Font(color=C_SUBTEXT, size=10, name="Calibri")
    c.alignment = align(h="left")
    ws4.row_dimensions[r4].height = 18
    r4 += 1

col_widths_4 = [16, 12, 14, 16, 10, 16, 16, 16]
for i, w in enumerate(col_widths_4, 1):
    set_col_width(ws4, i, w)

# ── Arka plan tamamen doldur ──────────────────────────────────────────────
for ws in [ws1, ws2, ws3, ws4]:
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", ""):
                cell.fill = fill(BG_DARK)

# ── Dosyayı kaydet ────────────────────────────────────────────────────────
out_path = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "nutrilens_fiyatlandirma_modeli.xlsx"
)
wb.save(out_path)
print(f"✅  Excel kaydedildi: {out_path}")
